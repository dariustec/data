
from openpyxl import load_workbook
from xlrd.timemachine import xrange

wb = load_workbook(filename='ap.xlsx')
ws = wb.active
search_words = ['PID']


def print_row(row):
    line = ''
    for col in xrange(1, ws.max_column + 1):
        _cell = ws.cell(row=row, column=col).value
        if _cell:
            line += ' ' + str(_cell)
    return line


for row in xrange(1, ws.max_row + 1):
    for col in xrange(1, ws.max_column + 1):
        _cell = ws.cell(row=row, column=col)
        if any(word in str(_cell.value) for word in search_words):
            print(print_row(row))
            break
